"""
Step 3A — export Portfolio Pulse Python analytics into the Excel reporting model.
"""
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from market_engine import run_current_analysis, run_stress_tests

DEFAULT_WORKBOOK = Path("excel/Portfolio_Pulse_Model.xlsx")

def _clean_date(value):
    if value is None:
        return "Not recovered by scenario end"
    try:
        if value != value:  # pandas NaT / NaN
            return "Not recovered by scenario end"
    except Exception:
        pass
    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime()
    return value

def export_to_excel(workbook_path=DEFAULT_WORKBOOK):
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(
            f"Workbook not found: {workbook_path}\n"
            "Put Portfolio_Pulse_Model.xlsx inside the excel/ folder."
        )

    current = run_current_analysis()
    stress = run_stress_tests()
    wb = load_workbook(workbook_path)

    for sheet in ("Analytics", "Stress Tests", "Dashboard"):
        if sheet not in wb.sheetnames:
            raise ValueError(f"Workbook is missing required sheet: {sheet}")

    analytics = wb["Analytics"]
    stress_ws = wb["Stress Tests"]
    dashboard = wb["Dashboard"]

    p = current["portfolio_stats"]
    b = current["benchmark_stats"]

    values = {
        "B4": p["annualized_return"], "C4": b["annualized_return"],
        "B5": p["annualized_volatility"], "C5": b["annualized_volatility"],
        "B6": p["sharpe_ratio"], "C6": b["sharpe_ratio"],
        "B7": p["max_drawdown"], "C7": b["max_drawdown"],
        "B8": p["beta"], "C8": 1.0,
        "B9": p["excess_annualized_return"], "C9": 0.0,
    }
    for cell, value in values.items():
        analytics[cell] = float(value)

    for cell in ("B4","C4","B5","C5","B7","C7","B9","C9"):
        analytics[cell].number_format = "0.00%"
    for cell in ("B6","C6","B8","C8"):
        analytics[cell].number_format = "0.00"

    # Audit trail
    analytics["A11"] = "Analysis Start"
    analytics["B11"] = current["prices"].index.min().to_pydatetime()
    analytics["A12"] = "Analysis End"
    analytics["B12"] = current["prices"].index.max().to_pydatetime()
    analytics["A13"] = "Last Python Refresh"
    analytics["B13"] = datetime.now()
    for cell in ("B11","B12"):
        analytics[cell].number_format = "yyyy-mm-dd"
    analytics["B13"].number_format = "yyyy-mm-dd hh:mm"

    # Historical stress-test table
    for excel_row, (_, row) in enumerate(stress.iterrows(), start=4):
        stress_ws.cell(excel_row, 1, row["scenario"])
        stress_ws.cell(excel_row, 2, _clean_date(row["start"]))
        stress_ws.cell(excel_row, 3, _clean_date(row["end"]))
        stress_ws.cell(excel_row, 4, float(row["period_return"]))
        stress_ws.cell(excel_row, 5, float(row["max_drawdown"]))
        stress_ws.cell(excel_row, 6, _clean_date(row["recovery_date"]))
        stress_ws.cell(excel_row, 4).number_format = "0.00%"
        stress_ws.cell(excel_row, 5).number_format = "0.00%"

    # Dashboard summary
    dashboard["D12"], dashboard["E12"] = "Portfolio CAGR", float(p["annualized_return"])
    dashboard["D13"], dashboard["E13"] = "60/40 CAGR", float(b["annualized_return"])
    dashboard["D14"], dashboard["E14"] = "Portfolio Max Drawdown", float(p["max_drawdown"])
    dashboard["D15"], dashboard["E15"] = "Sharpe Ratio", float(p["sharpe_ratio"])
    dashboard["D16"], dashboard["E16"] = "Excess Annualized Return", float(p["excess_annualized_return"])
    for cell in ("E12","E13","E14","E16"):
        dashboard[cell].number_format = "0.00%"
    dashboard["E15"].number_format = "0.00"

    # Never overwrite the template.
    output = workbook_path.with_name("Portfolio_Pulse_Analysis.xlsx")
    wb.save(output)
    return output

if __name__ == "__main__":
    output = export_to_excel()
    print("\nPORTFOLIO PULSE — EXCEL EXPORT")
    print("=" * 50)
    print(f"Created: {output.resolve()}")
