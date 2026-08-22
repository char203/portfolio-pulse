"""
Portfolio Pulse — Excel integration with:
- wealth curve
- attribution
- allocation sensitivity analysis

Replace your current excel_bidirectional.py with this file.
"""
from pathlib import Path
from datetime import datetime
import sys
import pandas as pd
from openpyxl import load_workbook

from market_engine import run_current_analysis, run_stress_tests
from attribution import (
    calculate_daily_contributions,
    reconcile_contributions,
    summarize_contributions,
)
from sensitivity import (
    build_sensitivity_scenarios,
    validate_scenarios,
    sensitivity_table,
)

DEFAULT_WORKBOOK = Path("excel/Portfolio_Pulse_Analysis.xlsx")
TEMPLATE_FALLBACK = Path("excel/Portfolio_Pulse_Model.xlsx")
WEIGHT_ROWS = {9: "VTI", 10: "VXUS", 11: "AGG", 12: "SGOV"}


def read_excel_inputs(workbook_path):
    keep_vba = workbook_path.suffix.lower() == ".xlsm"
    wb = load_workbook(workbook_path, data_only=False, keep_vba=keep_vba)
    ws = wb["Portfolio"]
    weights = {}
    for row, expected in WEIGHT_ROWS.items():
        ticker = str(ws.cell(row, 2).value).strip().upper()
        weight = ws.cell(row, 3).value
        if ticker != expected:
            raise ValueError(f"Portfolio row {row}: expected {expected}, found {ticker}.")
        if not isinstance(weight, (int, float)):
            raise ValueError(f"{ticker} weight must be numeric.")
        weights[ticker] = float(weight)

    return wb, {
        "objective": ws["B3"].value,
        "time_horizon": ws["B4"].value,
        "risk_tolerance": ws["B5"].value,
        "portfolio_value": ws["B6"].value,
        "weights": weights,
    }


def validate_inputs(inputs):
    total = sum(inputs["weights"].values())
    if abs(total - 1.0) > 0.0001:
        raise ValueError(f"Weights must sum to 100%; current total is {total:.2%}.")
    for ticker, weight in inputs["weights"].items():
        if weight < 0 or weight > 1:
            raise ValueError(f"Invalid {ticker} weight: {weight:.2%}.")
    if not isinstance(inputs["portfolio_value"], (int, float)) or inputs["portfolio_value"] <= 0:
        raise ValueError("Portfolio value must be greater than zero.")
    if not isinstance(inputs["risk_tolerance"], (int, float)) or not 1 <= inputs["risk_tolerance"] <= 10:
        raise ValueError("Risk tolerance must be between 1 and 10.")


def write_analytics(wb, inputs, current):
    ws = wb["Analytics"]
    p, b = current["portfolio_stats"], current["benchmark_stats"]
    values = {
        "B4": p["annualized_return"], "C4": b["annualized_return"],
        "B5": p["annualized_volatility"], "C5": b["annualized_volatility"],
        "B6": p["sharpe_ratio"], "C6": b["sharpe_ratio"],
        "B7": p["max_drawdown"], "C7": b["max_drawdown"],
        "B8": p["beta"], "C8": 1.0,
        "B9": p["excess_annualized_return"], "C9": 0.0,
    }
    for cell, value in values.items():
        ws[cell] = float(value)
    for cell in ("B4","C4","B5","C5","B7","C7","B9","C9"):
        ws[cell].number_format = "0.00%"
    for cell in ("B6","C6","B8","C8"):
        ws[cell].number_format = "0.00"

    ws["A11"], ws["B11"] = "Analysis Start", current["prices"].index.min().to_pydatetime()
    ws["A12"], ws["B12"] = "Analysis End", current["prices"].index.max().to_pydatetime()
    ws["A13"], ws["B13"] = "Last Python Refresh", datetime.now()
    ws["A14"] = "Portfolio Weights Used"
    ws["B14"] = ", ".join(f"{k} {v:.1%}" for k, v in inputs["weights"].items())
    ws["A16"], ws["B16"] = "Ending Portfolio Value", float(current["portfolio_wealth"].iloc[-1])
    ws["A17"], ws["B17"] = "Ending 60/40 Value", float(current["benchmark_wealth"].iloc[-1])
    ws["B11"].number_format = ws["B12"].number_format = "yyyy-mm-dd"
    ws["B13"].number_format = "yyyy-mm-dd hh:mm"
    ws["B16"].number_format = ws["B17"].number_format = "$#,##0.00"


def write_stress_tests(wb, stress):
    ws = wb["Stress Tests"]
    for r, (_, row) in enumerate(stress.iterrows(), start=4):
        ws.cell(r,1,row["scenario"]); ws.cell(r,2,row["start"]); ws.cell(r,3,row["end"])
        ws.cell(r,4,float(row["period_return"])); ws.cell(r,5,float(row["max_drawdown"]))
        recovery = row["recovery_date"]
        ws.cell(r,6, "Not recovered by scenario end" if recovery != recovery else recovery)
        ws.cell(r,4).number_format = ws.cell(r,5).number_format = "0.00%"


def write_wealth_curve(wb, current):
    name = "Wealth Curve Data"
    ws = wb[name] if name in wb.sheetnames else wb.create_sheet(name)
    ws.delete_rows(1, ws.max_row)
    ws.append(["Date","Portfolio Pulse","60/40 Benchmark","Dollar Difference","Relative Difference"])
    combined = current["portfolio_wealth"].rename("portfolio").to_frame().join(
        current["benchmark_wealth"].rename("benchmark").to_frame(), how="inner")
    for date, row in combined.iterrows():
        p, b = float(row["portfolio"]), float(row["benchmark"])
        ws.append([date.to_pydatetime(), p, b, p-b, p/b-1 if b else None])
    for r in range(2, ws.max_row+1):
        ws.cell(r,1).number_format = "yyyy-mm-dd"
        for c in (2,3,4): ws.cell(r,c).number_format = "$#,##0.00"
        ws.cell(r,5).number_format = "0.00%"


def write_attribution(wb, inputs, current):
    prices = current["prices"][list(inputs["weights"].keys())]
    asset_returns = prices.pct_change().dropna()
    contributions = calculate_daily_contributions(asset_returns, inputs["weights"])
    portfolio_returns = asset_returns.mul(pd.Series(inputs["weights"]), axis="columns").sum(axis=1)

    if not reconcile_contributions(contributions, portfolio_returns):
        raise ValueError("Attribution reconciliation failed.")

    summary = summarize_contributions(contributions)
    name = "Attribution"
    ws = wb[name] if name in wb.sheetnames else wb.create_sheet(name)
    ws.delete_rows(1, ws.max_row)
    ws.append(["Asset","Weight","Cumulative Arithmetic Contribution","Rank"])
    for rank, (ticker, contribution) in enumerate(summary.items(), start=1):
        ws.append([ticker, inputs["weights"][ticker], float(contribution), rank])

    ws.append([])
    ws.append(["Control","Value"])
    ws.append(["Daily reconciliation","PASS"])
    ws.append(["Sum of arithmetic contributions",float(summary.sum())])
    ws.append(["Note","Sum of daily weight × return; not Brinson attribution."])

    for r in range(2, 2+len(summary)):
        ws.cell(r,2).number_format = "0.0%"
        ws.cell(r,3).number_format = "0.00%"


def write_sensitivity(wb, inputs):
    scenarios = build_sensitivity_scenarios(inputs["weights"])
    validate_scenarios(scenarios)

    results = {}
    for name, weights in scenarios.items():
        results[name] = run_current_analysis(
            weights=weights,
            initial_value=float(inputs["portfolio_value"]),
        )

    table = sensitivity_table(results)

    name = "Sensitivity"
    ws = wb[name] if name in wb.sheetnames else wb.create_sheet(name)
    ws.delete_rows(1, ws.max_row)

    headers = ["Scenario","CAGR","Volatility","Sharpe","Max Drawdown","Ending Value"]
    ws.append(headers)

    for _, row in table.iterrows():
        ws.append([
            row["Scenario"],
            float(row["CAGR"]),
            float(row["Volatility"]),
            float(row["Sharpe"]),
            float(row["Max Drawdown"]),
            float(row["Ending Value"]),
        ])

    for r in range(2, ws.max_row+1):
        ws.cell(r,2).number_format = "0.00%"
        ws.cell(r,3).number_format = "0.00%"
        ws.cell(r,4).number_format = "0.00"
        ws.cell(r,5).number_format = "0.00%"
        ws.cell(r,6).number_format = "$#,##0.00"

    ws["H1"] = "Interpretation"
    ws["H2"] = "Illustrative allocation shifts around the selected base portfolio."
    ws["H3"] = "These are sensitivity scenarios, not recommendations."


def main():
    workbook_path = (
        Path(sys.argv[1]).expanduser().resolve()
        if len(sys.argv) > 1
        else (DEFAULT_WORKBOOK if DEFAULT_WORKBOOK.exists() else TEMPLATE_FALLBACK)
    )
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    print("\nPORTFOLIO PULSE — SENSITIVITY RUN")
    print("=" * 55)
    print(f"Reading inputs from: {workbook_path}")

    wb, inputs = read_excel_inputs(workbook_path)
    validate_inputs(inputs)

    current = run_current_analysis(
        weights=inputs["weights"],
        initial_value=float(inputs["portfolio_value"]),
    )
    stress = run_stress_tests(weights=inputs["weights"])

    write_analytics(wb, inputs, current)
    write_stress_tests(wb, stress)
    write_wealth_curve(wb, current)
    write_attribution(wb, inputs, current)
    write_sensitivity(wb, inputs)

    wb.save(workbook_path)

    print("\nAnalysis complete.")
    print(f"Updated: {workbook_path}")
    print("Attribution reconciliation: PASS")
    print("Created/updated sheet: Sensitivity")


if __name__ == "__main__":
    main()
